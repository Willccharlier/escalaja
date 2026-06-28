import json
from calendar import monthrange
from datetime import datetime, date, time, timedelta

# Django
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.utils import timezone
from django.views.decorators.http import require_POST

# Bibliotecas de terceiros
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# Imports locais
from .models import Escala, DiaEscala, Funcionario, Turno, Feriado, ConfiguracaoSistema, Grupo, SetorTurno
from .services import GeradorEscala



def login_view(request):
    """View de login personalizada"""
    if request.user.is_authenticated:
        return redirect('escala:dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            messages.success(request, f'✅ Bem-vindo, {user.get_full_name() or user.username}!')
            next_url = request.GET.get('next', 'escala:dashboard')
            return redirect(next_url)
        else:
            messages.error(request, '❌ Usuário ou senha incorretos!')
    
    return render(request, 'escala/login.html')


def logout_view(request):
    """View de logout"""
    logout(request)
    messages.info(request, '👋 Você saiu do sistema com sucesso!')
    return redirect('escala:login')


# ==================== DASHBOARD ====================

@login_required
def dashboard(request):
    agora = timezone.localtime()
    hoje = agora.date()
    hora_atual = agora.time()
    tz = timezone.get_current_timezone()

    # ===== TURNOS =====
    turnos = list(Turno.objects.all().order_by('horario_entrada'))
    turno_atual = turno_anterior = turno_proximo = None

    for i, t in enumerate(turnos):
        ini, fim = t.horario_entrada, t.horario_saida
        if (ini < fim and ini <= hora_atual < fim) or (ini > fim and (hora_atual >= ini or hora_atual < fim)):
            turno_atual = t
            turno_anterior = turnos[i - 1]
            turno_proximo = turnos[(i + 1) % len(turnos)]
            break

    if not turno_atual:
        return render(request, 'escala/dashboard.html', {})

    # ===== DATETIME REAL DO TURNO (CORRETO) =====
    inicio_data = hoje
    if turno_atual.horario_entrada > turno_atual.horario_saida and hora_atual < turno_atual.horario_saida:
        inicio_data = hoje - timedelta(days=1)

    inicio_dt = timezone.make_aware(
        datetime.combine(inicio_data, turno_atual.horario_entrada), tz
    )

    fim_data = inicio_data
    if turno_atual.horario_saida <= turno_atual.horario_entrada:
        fim_data = inicio_data + timedelta(days=1)

    fim_dt = timezone.make_aware(
        datetime.combine(fim_data, turno_atual.horario_saida), tz
    )

    # ===== PROGRESSO (NÃO COMEÇA EM 100%) =====
    total_segundos = max(1, (fim_dt - inicio_dt).total_seconds())
    decorridos = (agora - inicio_dt).total_seconds()
    progresso_turno = max(0, min(100, (decorridos / total_segundos) * 100))

    # ===== TEMPO RESTANTE DO ATUAL =====
    restante = max(0, int((fim_dt - agora).total_seconds()))
    horas = restante // 3600
    minutos = (restante % 3600) // 60
    tempo_restante = f"{horas:02d}:{minutos:02d}"

    # ===== TEMPO PARA O PRÓXIMO COMEÇAR =====
    proximo_inicio_dt = fim_dt  # próximo começa quando o atual termina
    seg_proximo = max(0, int((proximo_inicio_dt - agora).total_seconds()))
    hp = seg_proximo // 3600
    mp = (seg_proximo % 3600) // 60
    tempo_para_proximo = f"{hp:02d}:{mp:02d}"

    # ===== FUNCIONÁRIOS =====
    # Inclui regulares do turno + folguistas cobrindo o turno (via turno_coberto)
    from django.db.models import Q

    def _funcionarios_no_turno(data_dia, turno):
        return DiaEscala.objects.filter(
            Q(funcionario__turno=turno) | Q(turno_coberto=turno),
            data=data_dia,
            funcionario__ativo=True,
            situacao='TRABALHA'
        ).select_related('funcionario')

    funcionarios_turno_atual = _funcionarios_no_turno(inicio_data, turno_atual)

    data_turno_anterior = inicio_data
    if turno_anterior.horario_saida < turno_anterior.horario_entrada:
        data_turno_anterior = inicio_data - timedelta(days=1)

    funcionarios_turno_anterior = _funcionarios_no_turno(data_turno_anterior, turno_anterior)
    funcionarios_turno_proximo = _funcionarios_no_turno(fim_data, turno_proximo)

    context = {
        'turno_atual': turno_atual,
        'turno_anterior': turno_anterior,
        'turno_proximo': turno_proximo,
        'funcionarios_turno_atual': funcionarios_turno_atual,
        'funcionarios_turno_anterior': funcionarios_turno_anterior,
        'funcionarios_turno_proximo': funcionarios_turno_proximo,
        'progresso_turno': round(progresso_turno, 1),
        'tempo_restante': tempo_restante,
        'tempo_para_proximo': tempo_para_proximo,
    }

    return render(request, 'escala/dashboard.html', context)


# ==================== ESCALAS ====================

@login_required
def escala_lista(request):
    """Lista todas as escalas"""
    escalas = Escala.objects.all().order_by('-ano', '-mes')
    return render(request, 'escala/escala_lista.html', {'escalas': escalas})


@login_required
def escala_detalhe(request, pk):
    """Mostra detalhes de uma escala em formato de tabela por turno"""
    escala = get_object_or_404(Escala, pk=pk)

    dias_semana_abrev = {0: 'SEG', 1: 'TER', 2: 'QUA', 3: 'QUI', 4: 'SEX', 5: 'SAB', 6: 'DOM'}

    dias_mes = monthrange(escala.ano, escala.mes)[1]
    calendario_dias = []
    for dia in range(1, dias_mes + 1):
        d = date(escala.ano, escala.mes, dia)
        calendario_dias.append({
            'dia': dia,
            'dia_semana_abrev': dias_semana_abrev[d.weekday()],
            'eh_domingo': d.weekday() == 6
        })

    dias_escala = DiaEscala.objects.filter(escala=escala).select_related(
        'funcionario', 'funcionario__turno', 'funcionario__grupo', 'turno_coberto', 'setor_coberto'
    )

    SIT_MAP = {
        'TRABALHA':          ('trabalha',       '✔'),
        'FOLGA':             ('folga',           'F'),
        'FOLGA_COMPENSADA':  ('folga-compensada','C'),
        'FALTA':             ('falta',           'FT'),
        'ATESTADO':          ('atestado',        'AF'),
        'FERIAS':            ('ferias',          'FB'),
        'FOLGA_ANIVERSARIO': ('folga-aniv',      '🎂'),
        'FOLGA_FERIADO':     ('folga-feriado',   '🎉'),
    }

    def montar_linha(func_data, dias_turno=None):
        linha = {'id': func_data['id'], 'nome': func_data['nome'], 'dias': []}
        for dia in range(1, dias_mes + 1):
            situacao = func_data['dias_situacao'][dia]
            classe, simbolo = SIT_MAP.get(situacao, ('', ''))
            if situacao == 'TRABALHA' and dias_turno is not None:
                simbolo = dias_turno[dia].upper() if dias_turno[dia] else '?'
            linha['dias'].append({'classe': classe, 'simbolo': simbolo, 'situacao': situacao})
        return linha

    # Setor sections (REGULAR employees grouped by Grupo/Setor)
    setores = list(Grupo.objects.filter(
        funcionario__tipo='REGULAR', funcionario__ativo=True
    ).distinct().order_by('nome'))
    turnos_data = []  # mantém nome turnos_data para compatibilidade com template
    for setor in setores:
        func_dict = {}
        for dia_obj in dias_escala:
            if dia_obj.funcionario.tipo != 'REGULAR':
                continue
            if dia_obj.funcionario.grupo_id != setor.id:
                continue
            fid = dia_obj.funcionario.id
            if fid not in func_dict:
                turno_func = dia_obj.funcionario.turno
                horario = (f"{turno_func.horario_entrada.strftime('%H:%M')}–{turno_func.horario_saida.strftime('%H:%M')}"
                           if turno_func else '')
                func_dict[fid] = {
                    'id': fid,
                    'nome': dia_obj.funcionario.nome,
                    'turno_nome': turno_func.nome if turno_func else '',
                    'horario': horario,
                    'dias_situacao': [''] * (dias_mes + 1),
                }
            func_dict[fid]['dias_situacao'][dia_obj.data.day] = dia_obj.situacao
        if func_dict:
            # Ordenar por turno (MANHA → INTERMEDIARIO → TARDE → NOITE) e depois nome
            ORDEM_TURNO = {'MANHA': 0, 'INTERMEDIARIO': 1, 'TARDE': 2, 'NOITE': 3}
            fds_ordenados = sorted(
                func_dict.values(),
                key=lambda fd: (ORDEM_TURNO.get(fd['turno_nome'].upper(), 99), fd['nome'])
            )
            # Para regulares: mostra abreviação do turno nas células de TRABALHA
            funcionarios_linhas = []
            for fd in fds_ordenados:
                abrev = fd['turno_nome'][:3].upper() if fd['turno_nome'] else None
                dias_label = [abrev] * (dias_mes + 1)  # mesmo turno todos os dias
                funcionarios_linhas.append(montar_linha(fd, dias_turno=dias_label))
            turnos_data.append({
                'nome': setor.nome,
                'horario': '',
                'funcionarios': funcionarios_linhas,
                'func_detalhes': list(func_dict.values()),
            })

    # Folguistas section
    folguista_dict = {}
    for dia_obj in dias_escala:
        if dia_obj.funcionario.tipo != 'FOLGUISTA':
            continue
        fid = dia_obj.funcionario.id
        if fid not in folguista_dict:
            folguista_dict[fid] = {
                'id': fid,
                'nome': dia_obj.funcionario.nome,
                'dias_situacao': [''] * (dias_mes + 1),
                'dias_label': [None] * (dias_mes + 1),   # texto exibido na célula
                'dias_turno_id': [None] * (dias_mes + 1),
                'dias_setor_id': [None] * (dias_mes + 1),
            }
        d = dia_obj.data.day
        folguista_dict[fid]['dias_situacao'][d] = dia_obj.situacao
        if dia_obj.setor_coberto and dia_obj.turno_coberto:
            label = f"{dia_obj.setor_coberto.nome[:4]}/{dia_obj.turno_coberto.nome[:3]}"
            folguista_dict[fid]['dias_label'][d] = label
            folguista_dict[fid]['dias_turno_id'][d] = dia_obj.turno_coberto.id
            folguista_dict[fid]['dias_setor_id'][d] = dia_obj.setor_coberto.id
        elif dia_obj.turno_coberto:
            folguista_dict[fid]['dias_label'][d] = dia_obj.turno_coberto.nome[:4]
            folguista_dict[fid]['dias_turno_id'][d] = dia_obj.turno_coberto.id

    folguistas_data = [
        montar_linha(fd, dias_turno=fd['dias_label'])
        for fd in folguista_dict.values()
    ]

    context = {
        'escala': escala,
        'calendario_dias': calendario_dias,
        'turnos_data': turnos_data,
        'folguistas_data': folguistas_data,
        'turnos_disponiveis': Turno.objects.all().order_by('horario_entrada'),
        'setor_turnos_disponiveis': SetorTurno.objects.select_related('setor', 'turno').order_by('setor__nome', 'turno__horario_entrada'),
    }
    return render(request, 'escala/escala_detalhe.html', context)


@login_required
def gerar_escala_view(request):
    """Gera uma nova escala"""
    if request.method == 'POST':
        mes = int(request.POST.get('mes'))
        ano = int(request.POST.get('ano'))
        force = request.POST.get('force') == 'on'
        
        escala_existente = Escala.objects.filter(mes=mes, ano=ano).first()
        
        if escala_existente and not force:
            messages.warning(request, f'Já existe escala para {mes:02d}/{ano}. Marque "Forçar regerar" para substituir.')
            return redirect('escala:gerar_escala')
        
        if escala_existente and force:
            escala_existente.delete()
            messages.info(request, f'Escala anterior de {mes:02d}/{ano} removida.')
        
        gerador = GeradorEscala(mes, ano)
        sucesso, escala, alertas = gerador.gerar()

        if escala is None:
            messages.error(request, '❌ Erro crítico ao gerar escala. Verifique os detalhes do sistema.')
            return redirect('escala:gerar_escala')

        if sucesso:
            messages.success(request, f'✅ Escala de {mes:02d}/{ano} gerada com sucesso!')
        else:
            messages.error(request, f'⚠️ Escala gerada com problemas. Abra "Ver detalhes do sistema" para mais informações.')

        return redirect('escala:escala_detalhe', pk=escala.id)
    
    hoje = date.today()
    return render(request, 'escala/gerar_escala.html', {
        'mes_atual': hoje.month,
        'ano_atual': hoje.year,
    })


@login_required
def revalidar_escala(request, pk):
    """Revalida uma escala — mostra todas as validações com status detalhado por funcionário/setor."""
    from calendar import monthrange
    from .models import SetorTurno

    escala = get_object_or_404(Escala, pk=pk)
    dias_mes = monthrange(escala.ano, escala.mes)[1]
    config = ConfiguracaoSistema.get()

    # Carregar escala em memória
    todos_dias = list(DiaEscala.objects.filter(escala=escala).select_related(
        'funcionario', 'funcionario__turno', 'funcionario__grupo',
        'setor_coberto', 'turno_coberto'
    ))
    escala_mem = {}   # {func_id: {dia: situacao}}
    setor_cob  = {}   # {func_id: {dia: setor_id}}
    turno_cob  = {}   # {func_id: {dia: turno_id}}
    for d in todos_dias:
        fid = d.funcionario.id
        escala_mem.setdefault(fid, {})[d.data.day] = d.situacao
        if d.funcionario.tipo == 'FOLGUISTA':
            if d.setor_coberto:
                setor_cob.setdefault(fid, {})[d.data.day] = d.setor_coberto.id
            if d.turno_coberto:
                turno_cob.setdefault(fid, {})[d.data.day] = d.turno_coberto.id

    funcionarios_map = {
        f.id: f for f in Funcionario.objects.filter(
            id__in=list(escala_mem.keys()), ativo=True
        ).select_related('grupo', 'turno')
    }

    domingos = [d for d in range(1, dias_mes + 1) if date(escala.ano, escala.mes, d).weekday() == 6]
    setor_turnos = list(SetorTurno.objects.select_related('setor', 'turno').all())

    alertas = []
    tem_erro = False

    # ── REGRA 1: Domingo garantido (R1 — obrigatório para todos) ─────────────
    alertas.append("═══ R1: DOMINGO GARANTIDO — 1 domingo de folga por mês ═══")
    erros_r1 = []
    for func_id, dias in escala_mem.items():
        func = funcionarios_map.get(func_id)
        if not func:
            continue
        dom = next((d for d in domingos if dias.get(d, 'TRABALHA') != 'TRABALHA'), None)
        if dom:
            alertas.append(f"   ✅ {func.nome} ({func.regime}): folga no domingo {dom:02d}/{escala.mes:02d}")
        else:
            erros_r1.append(f"   ❌ {func.nome} ({func.regime}): SEM DOMINGO DE FOLGA")
            tem_erro = True
    if erros_r1:
        alertas.extend(erros_r1)
    alertas.append(f"   {'✅ Todos com domingo garantido!' if not erros_r1 else f'❌ {len(erros_r1)} funcionário(s) sem domingo!'}")

    # ── REGRA 2: Quantidade de folgas no mês (R2 CLT) ────────────────────────
    alertas.append("\n═══ R2: FOLGAS MENSAIS — quantidade correta por regime ═══")

    def semanas_do_mes():
        """Divide o mês em semanas de domingo a sábado — idêntico ao services.py."""
        semanas, semana_atual = [], []
        for dia in range(1, dias_mes + 1):
            wd = date(escala.ano, escala.mes, dia).weekday()
            if wd == 6 and semana_atual:   # domingo inicia nova semana
                semanas.append(semana_atual)
                semana_atual = []
            semana_atual.append(dia)
            if wd == 5 or dia == dias_mes:  # sábado ou último dia fecha semana
                semanas.append(semana_atual)
                semana_atual = []
        if semana_atual:
            semanas.append(semana_atual)
        return semanas

    def folgas_esperadas(func):
        semanas = semanas_do_mes()
        total = 0
        for s in semanas:
            n = len(s)
            if func.regime == '5x2':
                total += 2 if n == 7 else (1 if n >= 3 else 0)
            else:  # 6x1
                total += 1 if n >= 4 else 0
        return total

    erros_r2 = []
    for func_id, dias in escala_mem.items():
        func = funcionarios_map.get(func_id)
        if not func:
            continue
        esperado = folgas_esperadas(func)
        real = sum(1 for s in dias.values() if s != 'TRABALHA')
        if real == esperado:
            alertas.append(f"   ✅ {func.nome} ({func.regime}): {real}/{esperado} folgas")
        else:
            erros_r2.append(f"   ❌ {func.nome} ({func.regime}): {real}/{esperado} folgas (esperado {esperado})")
            tem_erro = True
    if erros_r2:
        alertas.extend(erros_r2)
    alertas.append(f"   {'✅ Todas as folgas mensais corretas!' if not erros_r2 else f'❌ {len(erros_r2)} funcionário(s) com folgas incorretas!'}")

    # ── REGRA 3: Máx consecutivos de trabalho (CLT) ───────────────────────────
    alertas.append("\n═══ R3: CONSECUTIVOS DE TRABALHO — limite por regime ═══")
    limite_consec = {'5x2': 5, '6x1': 6}
    erros_r3 = []
    for func_id, dias in escala_mem.items():
        func = funcionarios_map.get(func_id)
        if not func:
            continue
        lim = limite_consec.get(func.regime, 6)
        maximo = consecutivos = 0
        for d in range(1, dias_mes + 1):
            if dias.get(d, 'TRABALHA') == 'TRABALHA':
                consecutivos += 1
                maximo = max(maximo, consecutivos)
            else:
                consecutivos = 0
        if maximo <= lim:
            alertas.append(f"   ✅ {func.nome} ({func.regime}): máx {maximo} dias consecutivos (limite {lim})")
        else:
            erros_r3.append(f"   ❌ {func.nome} ({func.regime}): {maximo} dias consecutivos — ACIMA do limite {lim}!")
            tem_erro = True
    if erros_r3:
        alertas.extend(erros_r3)
    alertas.append(f"   {'✅ Nenhuma violação de consecutivos!' if not erros_r3 else f'❌ {len(erros_r3)} funcionário(s) com consecutivos acima do limite!'}")

    # ── REGRA 4: Folgas consecutivas (configurável) ───────────────────────────
    alertas.append("\n═══ R4: FOLGAS CONSECUTIVAS (configurável) ═══")
    if not config.consecutivas_ativo:
        alertas.append("   ⏭️ Regra desativada nas configurações")
    else:
        erros_r4 = []
        for func_id, dias in escala_mem.items():
            func = funcionarios_map.get(func_id)
            if not func:
                continue
            aplica = config.consecutivas_regime == 'AMBOS' or config.consecutivas_regime == func.regime
            if not aplica:
                continue
            for d in range(1, dias_mes):
                if dias.get(d) == 'FOLGA' and dias.get(d + 1) == 'FOLGA':
                    erros_r4.append(f"   ❌ {func.nome}: folgas consecutivas dias {d} e {d+1}")
                    tem_erro = True
                    break
            else:
                alertas.append(f"   ✅ {func.nome}: sem folgas consecutivas")
        if erros_r4:
            alertas.extend(erros_r4)
        alertas.append(f"   {'✅ Sem folgas consecutivas!' if not erros_r4 else f'❌ {len(erros_r4)} funcionário(s) com folgas consecutivas!'}")

    # ── REGRA 5: Lotação mínima ───────────────────────────────────────────────
    alertas.append("\n═══ R5: LOTAÇÃO MÍNIMA POR SETOR/TURNO ═══")
    erros_r5 = []
    for st in setor_turnos:
        funcs_st = list(Funcionario.objects.filter(tipo='REGULAR', ativo=True, grupo=st.setor, turno=st.turno))
        if not funcs_st:
            continue
        dias_com_erro = []
        for dia in range(1, dias_mes + 1):
            reg = sum(1 for f in funcs_st if escala_mem.get(f.id, {}).get(dia, 'TRABALHA') == 'TRABALHA')
            folg = sum(
                1 for fid in setor_cob
                if setor_cob[fid].get(dia) == st.setor.id
                and turno_cob.get(fid, {}).get(dia) == st.turno.id
                and escala_mem.get(fid, {}).get(dia) == 'TRABALHA'
            )
            total = reg + folg
            if total < st.minimo_funcionarios and not st.permite_zero:
                dias_com_erro.append(f"dia {dia:02d} ({total}/{st.minimo_funcionarios})")
        if dias_com_erro:
            erros_r5.append(f"   ❌ {st.setor.nome}/{st.turno.nome}: problemas em {', '.join(dias_com_erro)}")
            tem_erro = True
        else:
            pz = " (permite zero)" if st.permite_zero else ""
            alertas.append(f"   ✅ {st.setor.nome}/{st.turno.nome}: lotação OK em todos os dias{pz}")
    if erros_r5:
        alertas.extend(erros_r5)
    alertas.append(f"   {'✅ Lotação mínima OK em todos os setores!' if not erros_r5 else f'❌ {len(erros_r5)} setor(es)/turno(s) com deficit!'}")

    # ── Resultado final ───────────────────────────────────────────────────────
    alertas.append("\n" + "═" * 50)
    if tem_erro:
        alertas.append("❌ ESCALA COM PROBLEMAS — verifique os itens marcados com ❌")
        escala.gerada_com_sucesso = False
    else:
        alertas.append("✅ ESCALA VÁLIDA — todas as regras aprovadas!")
        escala.gerada_com_sucesso = True

    escala.observacoes = "\n".join(alertas)
    escala.save()

    messages.success(request, '🔄 Escala revalidada com sucesso!')
    return redirect('escala:escala_detalhe', pk=escala.id)


@login_required
def auto_corrigir_escala(request, pk):
    """Tenta corrigir automaticamente problemas de lotação mínima movendo folgas entre dias."""
    from calendar import monthrange
    from .models import SetorTurno
    from django.db import transaction

    escala = get_object_or_404(Escala, pk=pk)
    dias_mes = monthrange(escala.ano, escala.mes)[1]
    config = ConfiguracaoSistema.get()

    # Carregar escala atual do banco em memória
    dias_escala_qs = list(DiaEscala.objects.filter(escala=escala).select_related('funcionario', 'funcionario__grupo', 'funcionario__turno'))
    escala_mem = {}  # {func_id: {dia: situacao}}
    for d in dias_escala_qs:
        escala_mem.setdefault(d.funcionario.id, {})[d.data.day] = d.situacao

    # Mapear domingos do mês e o único domingo de folga de cada funcionário (R1 — intocável)
    domingos_mes = {d for d in range(1, dias_mes + 1) if date(escala.ano, escala.mes, d).weekday() == 6}

    def domingo_folga_de(func_id):
        """Retorna o dia de domingo que este funcionário tem de folga (None se não tiver)."""
        return next((d for d in domingos_mes if escala_mem.get(func_id, {}).get(d, 'TRABALHA') != 'TRABALHA'), None)

    def tem_domingo_apos_troca(func_id, dia_removido, dia_adicionado):
        """Verifica se o funcionário ainda terá domingo de folga após a troca."""
        dom = domingo_folga_de(func_id)
        if dom is None:
            return False  # Já não tem domingo — não piorar
        if dia_removido in domingos_mes and dia_removido == dom:
            # Estamos removendo o único domingo de folga
            # Só permitido se dia_adicionado também for domingo
            return dia_adicionado in domingos_mes
        return True

    setor_turnos = list(SetorTurno.objects.select_related('setor', 'turno').all())
    regulares_por_st = {}
    for st in setor_turnos:
        regulares_por_st[(st.setor.id, st.turno.id)] = list(
            Funcionario.objects.filter(tipo='REGULAR', ativo=True, grupo=st.setor, turno=st.turno)
        )

    def trabalhando_no_dia(st, dia):
        funcs = regulares_por_st.get((st.setor.id, st.turno.id), [])
        return sum(1 for f in funcs if escala_mem.get(f.id, {}).get(dia, 'TRABALHA') == 'TRABALHA')

    def tem_folgas_consecutivas(func_id):
        dias = escala_mem.get(func_id, {})
        for d in range(1, dias_mes):
            if dias.get(d) == 'FOLGA' and dias.get(d + 1) == 'FOLGA':
                return True
        return False

    def max_consecutivos_trabalho(func_id):
        dias = escala_mem.get(func_id, {})
        maximo = consecutivos = 0
        for d in range(1, dias_mes + 1):
            if dias.get(d, 'TRABALHA') == 'TRABALHA':
                consecutivos += 1
                maximo = max(maximo, consecutivos)
            else:
                consecutivos = 0
        return maximo

    limite_consec = {'5x2': 5, '6x1': 6}

    correcoes = 0
    relatorio = []

    # Carregar folguistas e coberturas logo no início (necessário para Fase 1b)
    folguistas = list(
        Funcionario.objects.filter(tipo='FOLGUISTA', ativo=True)
        .prefetch_related('grupos_habilitados', 'turnos_habilitados')
    )
    hab_setor = {f.id: set(f.grupos_habilitados.values_list('id', flat=True)) for f in folguistas}
    hab_turno = {f.id: set(f.turnos_habilitados.values_list('id', flat=True)) for f in folguistas}

    cobertura_folg = {}   # {func_id: {dia: (setor_id, turno_id)}}
    for d in DiaEscala.objects.filter(escala=escala, funcionario__tipo='FOLGUISTA').select_related('funcionario', 'setor_coberto', 'turno_coberto'):
        if d.setor_coberto and d.turno_coberto:
            cobertura_folg.setdefault(d.funcionario.id, {})[d.data.day] = (d.setor_coberto.id, d.turno_coberto.id)

    for d in DiaEscala.objects.filter(escala=escala, funcionario__tipo='FOLGUISTA').select_related('funcionario'):
        fid = d.funcionario.id
        escala_mem.setdefault(fid, {})[d.data.day] = d.situacao

    mudancas_cobertura = {}  # {(func_id, dia): (setor_id, turno_id)} — novas coberturas a salvar

    def folg_trabalhando_no_dia(st, dia):
        return sum(
            1 for fid, cobs in cobertura_folg.items()
            if cobs.get(dia) == (st.setor.id, st.turno.id)
            and escala_mem.get(fid, {}).get(dia) == 'TRABALHA'
        )

    # Mapa completo de funcionários (regulares + folguistas) para consulta de regime
    todos_funcionarios = {f.id: f for funcs in regulares_por_st.values() for f in funcs}
    todos_funcionarios.update({f.id: f for f in folguistas})

    # --- FASE 0: corrigir consecutivos de trabalho (CLT) ---
    for func_id, dias in escala_mem.items():
        func = todos_funcionarios.get(func_id)
        if not func:
            continue
        lim = limite_consec.get(func.regime, 6)

        for _ in range(30):
            # Encontrar primeira sequência que viola o limite
            seq, violacao = [], None
            for d in range(1, dias_mes + 1):
                if dias.get(d, 'TRABALHA') == 'TRABALHA':
                    seq.append(d)
                    if len(seq) > lim:
                        violacao = seq[:]
                        break
                else:
                    seq = []
            if not violacao:
                break

            # Tentar TROCAR: pegar folga de dia seguro e colocar no meio da violação
            meio = violacao[len(violacao) // 2]

            # Dias candidatos para "receber" a folga (meio da violação e vizinhos)
            candidatos_destino = sorted(violacao, key=lambda d: abs(d - meio))

            # Dias candidatos para "ceder" a folga (dias com folga, não domingo R1)
            dom_r1 = domingo_folga_de(func_id)
            candidatos_origem = [
                d for d, s in dias.items()
                if s == 'FOLGA' and d not in domingos_mes
                and d != dom_r1
                and d not in violacao
            ]
            # Preferir dias com mais cobertura de colegas (menos impacto ao devolver)
            def cob_regular(dia):
                st_func = next((st for st in setor_turnos
                                if st.setor.id == getattr(func, 'grupo_id', None)
                                and st.turno.id == getattr(func, 'turno_id', None)), None)
                if not st_func:
                    return 0
                return trabalhando_no_dia(st_func, dia)

            candidatos_origem.sort(key=cob_regular, reverse=True)

            trocado = False
            for dia_orig in candidatos_origem:
                for dia_dest in candidatos_destino:
                    if dia_dest == dia_orig:
                        continue
                    # Verificar se destino não cria deficit de lotação
                    st_func = next((st for st in setor_turnos
                                    if st.setor.id == getattr(func, 'grupo_id', None)
                                    and st.turno.id == getattr(func, 'turno_id', None)), None)
                    if st_func and not st_func.permite_zero:
                        trab_dest = trabalhando_no_dia(st_func, dia_dest)
                        if trab_dest < st_func.minimo_funcionarios:
                            continue  # Criaria deficit
                    # Fazer troca
                    dias[dia_orig] = 'TRABALHA'
                    dias[dia_dest] = 'FOLGA'
                    # Verificar se resolveu sem criar nova violação
                    nova_seq, nova_viol = [], None
                    for d2 in range(1, dias_mes + 1):
                        if dias.get(d2, 'TRABALHA') == 'TRABALHA':
                            nova_seq.append(d2)
                            if len(nova_seq) > lim:
                                nova_viol = True
                                break
                        else:
                            nova_seq = []
                    if not nova_viol:
                        relatorio.append(f"✅ {func.nome}: consecutivos corrigidos (folga {dia_orig:02d}→{dia_dest:02d})")
                        correcoes += 1
                        trocado = True
                        break
                    else:
                        dias[dia_orig] = 'FOLGA'
                        dias[dia_dest] = 'TRABALHA'
                if trocado:
                    break

            if not trocado:
                break  # Não conseguiu corrigir — para para este funcionário

    def tentar_mover_folga(func_id, dia_folga, dia_destino, st):
        """Testa mover FOLGA de dia_folga para dia_destino. Aplica em escala_mem se válido. Retorna True se funcionou."""
        func_regime = next((f.regime for funcs in regulares_por_st.values() for f in funcs if f.id == func_id), '6x1')
        lim = limite_consec.get(func_regime, 6)

        escala_mem[func_id][dia_folga] = 'TRABALHA'
        escala_mem[func_id][dia_destino] = 'FOLGA'

        valido = True
        if tem_folgas_consecutivas(func_id):
            valido = False
        if max_consecutivos_trabalho(func_id) > lim:
            valido = False
        # Destino não pode criar déficit neste mesmo setor/turno
        if valido and trabalhando_no_dia(st, dia_destino) < st.minimo_funcionarios and not st.permite_zero:
            valido = False

        if not valido:
            escala_mem[func_id][dia_folga] = 'FOLGA'
            escala_mem[func_id][dia_destino] = 'TRABALHA'

        return valido

    # Loop: revalida e corrige até não restar problemas ou estar genuinamente preso
    problemas_anteriores = None
    while True:
        problemas = []
        for dia in range(1, dias_mes + 1):
            for st in setor_turnos:
                if st.permite_zero or not regulares_por_st.get((st.setor.id, st.turno.id)):
                    continue
                if trabalhando_no_dia(st, dia) < st.minimo_funcionarios:
                    problemas.append((dia, st))

        if not problemas:
            break  # Tudo resolvido

        # Se os problemas são exatamente os mesmos da passagem anterior, estamos presos
        chave_problemas = frozenset((d, st.setor.id, st.turno.id) for d, st in problemas)
        if chave_problemas == problemas_anteriores:
            break  # Sem progresso — para
        problemas_anteriores = chave_problemas

        for dia_prob, st in problemas:
            if trabalhando_no_dia(st, dia_prob) >= st.minimo_funcionarios:
                continue  # Já foi resolvido numa iteração anterior deste loop

            funcs = regulares_por_st.get((st.setor.id, st.turno.id), [])
            de_folga = [f for f in funcs
                        if escala_mem.get(f.id, {}).get(dia_prob, 'TRABALHA') == 'FOLGA'
                        and not (dia_prob in domingos_mes and dia_prob == domingo_folga_de(f.id))]

            for func in de_folga:
                func_id = func.id
                # Deslizar a folga: +1, +2... até fim do mês, depois -1, -2...
                candidatos = (
                    list(range(dia_prob + 1, dias_mes + 1)) +
                    list(range(dia_prob - 1, 0, -1))
                )
                for dia_destino in candidatos:
                    if dia_destino in domingos_mes:
                        continue
                    if escala_mem.get(func_id, {}).get(dia_destino, 'TRABALHA') != 'TRABALHA':
                        continue
                    if tentar_mover_folga(func_id, dia_prob, dia_destino, st):
                        # escala_mem já foi atualizado em tentar_mover_folga — salva no final
                        relatorio.append(
                            f"✅ {func.nome}: folga {dia_prob:02d}→{dia_destino:02d}/{escala.mes:02d} "
                            f"({st.setor.nome}/{st.turno.nome})"
                        )
                        correcoes += 1
                        break

                if trabalhando_no_dia(st, dia_prob) >= st.minimo_funcionarios:
                    break  # Este problema foi resolvido, próximo

    # --- FASE 1b: deficit num domingo causado pelo domingo R1 do folguista → trocar domingo ---
    # Lógica: se folguista tem domingo off num dia com deficit, mover o domingo dele
    # para outro domingo do mês onde todos seus setores habilitados ficam cobertos sem ele.
    for dom_deficit in sorted(domingos_mes):
        for st in setor_turnos:
            if st.permite_zero:
                continue
            reg = trabalhando_no_dia(st, dom_deficit)
            folg = folg_trabalhando_no_dia(st, dom_deficit)
            if reg + folg >= st.minimo_funcionarios:
                continue  # Sem deficit aqui

            for func in folguistas:
                if st.setor.id not in hab_setor.get(func.id, set()):
                    continue
                if st.turno.id not in hab_turno.get(func.id, set()):
                    continue
                if escala_mem.get(func.id, {}).get(dom_deficit, 'TRABALHA') != 'FOLGA':
                    continue
                if domingo_folga_de(func.id) != dom_deficit:
                    continue  # Não é o domingo R1 dele

                # Tentar mover o domingo deste folguista para outro domingo do mês
                for novo_dom in sorted(domingos_mes - {dom_deficit}):
                    if escala_mem.get(func.id, {}).get(novo_dom, 'TRABALHA') != 'TRABALHA':
                        continue

                    # No novo domingo, todos os setores habilitados devem estar cobertos sem este folguista
                    novo_seguro = True
                    for st2 in setor_turnos:
                        if st2.permite_zero:
                            continue
                        if st2.setor.id not in hab_setor.get(func.id, set()):
                            continue
                        if st2.turno.id not in hab_turno.get(func.id, set()):
                            continue
                        reg2 = trabalhando_no_dia(st2, novo_dom)
                        folg2 = sum(
                            1 for fid, cobs in cobertura_folg.items()
                            if fid != func.id
                            and cobs.get(novo_dom) == (st2.setor.id, st2.turno.id)
                            and escala_mem.get(fid, {}).get(novo_dom) == 'TRABALHA'
                        )
                        if reg2 + folg2 < st2.minimo_funcionarios:
                            novo_seguro = False
                            break

                    if not novo_seguro:
                        continue

                    # Troca: trabalha no domingo com deficit, folga no domingo seguro
                    escala_mem[func.id][dom_deficit] = 'TRABALHA'
                    escala_mem[func.id][novo_dom] = 'FOLGA'
                    cobertura_folg.setdefault(func.id, {})[dom_deficit] = (st.setor.id, st.turno.id)
                    mudancas_cobertura[(func.id, dom_deficit)] = (st.setor.id, st.turno.id)
                    relatorio.append(
                        f"✅ {func.nome}: domingo {dom_deficit:02d}→{novo_dom:02d}/{escala.mes:02d} "
                        f"(cobrindo {st.setor.nome}/{st.turno.nome} dia {dom_deficit:02d})"
                    )
                    correcoes += 1
                    break

                if reg + folg_trabalhando_no_dia(st, dom_deficit) >= st.minimo_funcionarios:
                    break  # Resolvido

    # --- FASE 2: déficits que regulares não resolveram → tentar folguistas habilitados ---
    for dia in range(1, dias_mes + 1):
        for st in setor_turnos:
            if st.permite_zero or not regulares_por_st.get((st.setor.id, st.turno.id)):
                continue
            reg = trabalhando_no_dia(st, dia)
            folg = folg_trabalhando_no_dia(st, dia)
            if reg + folg >= st.minimo_funcionarios:
                continue  # OK

            # Procurar folguista habilitado
            for func in folguistas:
                if st.setor.id not in hab_setor.get(func.id, set()):
                    continue
                if st.turno.id not in hab_turno.get(func.id, set()):
                    continue

                sit = escala_mem.get(func.id, {}).get(dia, 'TRABALHA')

                if sit == 'TRABALHA':
                    # Já trabalhando — só reatribuir cobertura
                    cobertura_folg.setdefault(func.id, {})[dia] = (st.setor.id, st.turno.id)
                    mudancas_cobertura[(func.id, dia)] = (st.setor.id, st.turno.id)
                    relatorio.append(f"✅ {func.nome} → {st.setor.nome}/{st.turno.nome} dia {dia:02d}")
                    correcoes += 1
                    break

                if sit == 'FOLGA':
                    # Nunca remover o domingo R1 do folguista
                    if dia in domingos_mes and domingo_folga_de(func.id) == dia:
                        continue  # Tentar próximo folguista
                    # Déficit sem permite_zero = folguista habilitado vai obrigatoriamente
                    escala_mem[func.id][dia] = 'TRABALHA'
                    cobertura_folg.setdefault(func.id, {})[dia] = (st.setor.id, st.turno.id)
                    mudancas_cobertura[(func.id, dia)] = (st.setor.id, st.turno.id)
                    relatorio.append(f"✅ {func.nome}: convocado dia {dia:02d} → {st.setor.nome}/{st.turno.nome}")
                    correcoes += 1
                    break

    # Salvar todas as alterações em uma única transação
    todos_func = {f.id: f for funcs in regulares_por_st.values() for f in funcs}
    todos_func.update({f.id: f for f in folguistas})

    with transaction.atomic():
        for func_id, dias in escala_mem.items():
            func = todos_func.get(func_id)
            if not func:
                continue
            for dia, situacao in dias.items():
                d = date(escala.ano, escala.mes, dia)
                DiaEscala.objects.filter(escala=escala, funcionario=func, data=d).update(situacao=situacao)

        # Salvar mudanças de cobertura dos folguistas
        for (func_id, dia), (setor_id, turno_id) in mudancas_cobertura.items():
            func = todos_func.get(func_id)
            if not func:
                continue
            d = date(escala.ano, escala.mes, dia)
            DiaEscala.objects.filter(escala=escala, funcionario=func, data=d).update(
                setor_coberto_id=setor_id, turno_coberto_id=turno_id
            )

        # Revalidar e salvar observações
        alertas = []
        if correcoes > 0:
            alertas.append(f"🔧 AUTO-CORREÇÃO: {correcoes} ajuste(s) realizado(s):")
            alertas.extend([f"   {r}" for r in relatorio])
        else:
            alertas.append("ℹ️ AUTO-CORREÇÃO: Nenhum ajuste possível sem violar outras regras.")

        restantes = []
        for dia in range(1, dias_mes + 1):
            for st in setor_turnos:
                if st.permite_zero or not regulares_por_st.get((st.setor.id, st.turno.id)):
                    continue
                t = trabalhando_no_dia(st, dia) + folg_trabalhando_no_dia(st, dia)
                if t < st.minimo_funcionarios:
                    restantes.append(f"⚠️ DIA {dia:02d}/{escala.mes:02d} - {st.setor.nome}/{st.turno.nome}: {t}/{st.minimo_funcionarios}")

        if restantes:
            alertas.append("\n⚠️ Problemas restantes (não foi possível corrigir automaticamente):")
            alertas.extend([f"   {r}" for r in restantes])
            escala.gerada_com_sucesso = False
        else:
            alertas.append("\n✅ Lotação mínima OK em todos os dias!")
            escala.gerada_com_sucesso = True

        obs_anterior = escala.observacoes or ''
        escala.observacoes = "\n".join(alertas) + "\n\n---\n" + obs_anterior
        escala.save()

    msg = f'🔧 Auto-correção concluída: {correcoes} ajuste(s).' if correcoes else 'ℹ️ Nenhum ajuste possível sem violar outras regras.'
    messages.success(request, msg)
    # Redirecionar para revalidar para atualizar status completo (R1 + folgas + lotação)
    return redirect('escala:revalidar_escala', pk=escala.id)


@login_required
@require_POST
def alterar_situacao_dia(request):
    """Altera manualmente a situação de um dia na escala"""
    try:
        data = json.loads(request.body)
        escala_id = data['escala_id']
        funcionario_id = data['funcionario_id']
        dia = int(data['dia'])
        nova_situacao = data['nova_situacao']

        SITUACOES_VALIDAS = {'TRABALHA', 'FOLGA', 'FOLGA_COMPENSADA', 'FALTA', 'ATESTADO', 'FERIAS'}
        if nova_situacao not in SITUACOES_VALIDAS:
            return JsonResponse({'sucesso': False, 'erro': 'Situação inválida!'})

        escala = Escala.objects.get(id=escala_id)
        funcionario = Funcionario.objects.get(id=funcionario_id)
        data_dia = date(escala.ano, escala.mes, dia)

        dia_escala, _ = DiaEscala.objects.get_or_create(
            escala=escala,
            funcionario=funcionario,
            data=data_dia,
            defaults={'situacao': nova_situacao}
        )
        dia_escala.situacao = nova_situacao
        dia_escala.save()

        return JsonResponse({'sucesso': True})

    except Exception as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)})


@login_required
@require_POST
def alterar_turno_coberto(request):
    """Altera o setor+turno coberto por um folguista em um dia específico"""
    try:
        data = json.loads(request.body)
        escala_id = data['escala_id']
        funcionario_id = data['funcionario_id']
        dia = int(data['dia'])
        turno_id = data.get('turno_id') or None
        setor_id = data.get('setor_id') or None

        escala = Escala.objects.get(id=escala_id)
        funcionario = Funcionario.objects.get(id=funcionario_id)

        if funcionario.tipo != 'FOLGUISTA':
            return JsonResponse({'sucesso': False, 'erro': 'Funcionário não é folguista!'})

        data_dia = date(escala.ano, escala.mes, dia)
        turno = Turno.objects.get(id=turno_id) if turno_id else None
        setor = Grupo.objects.get(id=setor_id) if setor_id else None

        dia_escala = DiaEscala.objects.get(escala=escala, funcionario=funcionario, data=data_dia)
        dia_escala.turno_coberto = turno
        dia_escala.setor_coberto = setor
        dia_escala.save()

        return JsonResponse({'sucesso': True})

    except Exception as e:
        return JsonResponse({'sucesso': False, 'erro': str(e)})


@login_required
@require_POST
def trocar_folga(request):
    """Troca uma folga de dia, validando todas as regras"""
    try:
        data = json.loads(request.body)
        escala_id = data['escala_id']
        funcionario_id = data['funcionario_id']
        dia_origem = int(data['dia_origem'])
        dia_destino = int(data['dia_destino'])
        
        escala = Escala.objects.get(id=escala_id)
        funcionario = Funcionario.objects.get(id=funcionario_id)
        
        data_origem = date(escala.ano, escala.mes, dia_origem)
        data_destino = date(escala.ano, escala.mes, dia_destino)
        
        dia_escala_origem = DiaEscala.objects.get(
            escala=escala,
            funcionario=funcionario,
            data=data_origem
        )
        
        dia_escala_destino = DiaEscala.objects.get(
            escala=escala,
            funcionario=funcionario,
            data=data_destino
        )
        
        if dia_escala_origem.situacao != 'FOLGA':
            return JsonResponse({
                'sucesso': False,
                'erro': 'Só pode mover <strong>folgas regulares</strong>! Aniversários e feriados são fixos.'
            })
        
        if dia_escala_destino.situacao != 'TRABALHA':
            return JsonResponse({
                'sucesso': False,
                'erro': 'O destino deve ser um <strong>dia de trabalho</strong>!'
            })
        
        if not _mesma_semana(data_origem, data_destino):
            return JsonResponse({
                'sucesso': False,
                'erro': 'Só pode trocar folgas <strong>dentro da mesma semana</strong>!'
            })
        
        erro_domingo = _valida_domingo_obrigatorio(funcionario, escala, dia_origem, dia_destino)
        if erro_domingo:
            return JsonResponse({
                'sucesso': False,
                'erro': erro_domingo
            })
        
        if _cria_consecutivas(funcionario, escala, dia_origem, dia_destino):
            return JsonResponse({
                'sucesso': False,
                'erro': 'A troca criaria <strong>folgas consecutivas</strong>, o que não é permitido!'
            })
        
        if not _mantem_lotacao_minima(funcionario.turno, escala, data_origem, data_destino):
            return JsonResponse({
                'sucesso': False,
                'erro': 'A troca quebraria a <strong>lotação mínima</strong> do turno!'
            })
        
        dia_escala_origem.situacao = 'TRABALHA'
        dia_escala_destino.situacao = 'FOLGA'
        
        dia_escala_origem.save()
        dia_escala_destino.save()
        
        return JsonResponse({
            'sucesso': True,
            'mensagem': f'✅ Folga movida do <strong>dia {dia_origem}</strong> para o <strong>dia {dia_destino}</strong> com sucesso!'
        })
        
    except DiaEscala.DoesNotExist:
        return JsonResponse({
            'sucesso': False,
            'erro': 'Dia da escala não encontrado no banco de dados!'
        })
    except Exception as e:
        return JsonResponse({
            'sucesso': False,
            'erro': f'Erro inesperado: {str(e)}'
        })


def _mesma_semana(data1, data2):
    """Verifica se duas datas estão na mesma semana"""
    return data1.isocalendar()[1] == data2.isocalendar()[1]


def _valida_domingo_obrigatorio(funcionario, escala, dia_origem, dia_destino):
    """Valida se o funcionário manterá pelo menos 1 domingo de folga"""
    data_origem = date(escala.ano, escala.mes, dia_origem)
    eh_domingo_origem = data_origem.weekday() == 6
    
    if not eh_domingo_origem:
        return None
    
    dias = DiaEscala.objects.filter(
        escala=escala,
        funcionario=funcionario
    )
    
    domingos_folga = 0
    for d in dias:
        eh_domingo = d.data.weekday() == 6
        
        if not eh_domingo:
            continue
        
        if d.data.day == dia_origem:
            situacao_simulada = 'TRABALHA'
        elif d.data.day == dia_destino:
            situacao_simulada = 'FOLGA'
        else:
            situacao_simulada = d.situacao
        
        if situacao_simulada != 'TRABALHA':
            domingos_folga += 1
    
    if domingos_folga < 1:
        return (
            '⚠️ <strong>Operação bloqueada!</strong><br><br>'
            'Todo funcionário deve ter <strong>pelo menos 1 domingo de folga</strong> por mês.<br>'
            'Esta troca removeria o único domingo de descanso do funcionário.'
        )
    
    return None


def _cria_consecutivas(funcionario, escala, dia_origem, dia_destino):
    """Verifica se troca cria folgas consecutivas"""
    dias = DiaEscala.objects.filter(
        escala=escala,
        funcionario=funcionario
    ).order_by('data')
    
    situacoes = {}
    for d in dias:
        dia_num = d.data.day
        if dia_num == dia_origem:
            situacoes[dia_num] = 'TRABALHA'
        elif dia_num == dia_destino:
            situacoes[dia_num] = 'FOLGA'
        else:
            situacoes[dia_num] = d.situacao
    
    dias_ordenados = sorted(situacoes.keys())
    for i in range(len(dias_ordenados) - 1):
        hoje = dias_ordenados[i]
        amanha = dias_ordenados[i + 1]
        
        if amanha - hoje == 1:
            if situacoes[hoje] != 'TRABALHA' and situacoes[amanha] != 'TRABALHA':
                return True
    
    return False


def _mantem_lotacao_minima(turno, escala, data_origem, data_destino):
    """Verifica se mantém lotação mínima nos dois dias (por setor×turno)"""
    from .models import SetorTurno
    setor_turnos = SetorTurno.objects.filter(turno=turno).select_related('setor')

    for data in [data_origem, data_destino]:
        for st in setor_turnos:
            trabalhando = DiaEscala.objects.filter(
                escala=escala,
                data=data,
                funcionario__turno=turno,
                funcionario__grupo=st.setor,
                situacao='TRABALHA'
            ).count()

            if data == data_origem:
                trabalhando += 1
            elif data == data_destino:
                trabalhando -= 1

            if trabalhando < st.minimo_funcionarios:
                return False

    return True


# ==================== FUNCIONÁRIOS ====================

@login_required
def funcionario_lista(request):
    """Lista todos os funcionários"""
    funcionarios = Funcionario.objects.all().select_related('turno').order_by('turno', 'tipo', 'nome')
    
    regulares = funcionarios.filter(tipo='REGULAR', ativo=True)
    folguistas = funcionarios.filter(tipo='FOLGUISTA', ativo=True)
    inativos = funcionarios.filter(ativo=False)
    
    context = {
        'regulares': regulares,
        'folguistas': folguistas,
        'inativos': inativos,
    }
    return render(request, 'escala/funcionario_lista.html', context)


@login_required
def funcionario_novo(request):
    """Cadastra um novo funcionário"""
    if request.method == 'POST':
        try:
            nome = request.POST.get('nome')
            data_nascimento = request.POST.get('data_nascimento')
            data_admissao = request.POST.get('data_admissao')
            tipo = request.POST.get('tipo')
            turno_id = request.POST.get('turno')
            turno = Turno.objects.get(id=turno_id) if turno_id else None
            grupo_id = request.POST.get('grupo')
            grupo = Grupo.objects.get(id=grupo_id) if grupo_id else None
            regime = request.POST.get('regime', '6x1')
            folga_fixa = request.POST.get('folga_fixa_dia')
            folga_fixa_dia = int(folga_fixa) if folga_fixa not in ('', None) else None
            ferias_inicio = request.POST.get('ferias_inicio') or None
            ferias_fim = request.POST.get('ferias_fim') or None

            funcionario = Funcionario.objects.create(
                nome=nome,
                data_nascimento=data_nascimento,
                data_admissao=data_admissao,
                tipo=tipo,
                turno=turno,
                grupo=grupo,
                regime=regime,
                folga_fixa_dia=folga_fixa_dia,
                ativo=True,
                ferias_inicio=ferias_inicio,
                ferias_fim=ferias_fim,
            )

            # Turnos e grupos habilitados (folguista)
            turnos_hab = request.POST.getlist('turnos_habilitados')
            grupos_hab = request.POST.getlist('grupos_habilitados')
            if turnos_hab:
                funcionario.turnos_habilitados.set(Turno.objects.filter(id__in=turnos_hab))
            if grupos_hab:
                funcionario.grupos_habilitados.set(Grupo.objects.filter(id__in=grupos_hab))

            messages.success(request, f'✅ Funcionário {funcionario.nome} cadastrado com sucesso!')
            return redirect('escala:funcionario_lista')

        except Exception as e:
            messages.error(request, f'❌ Erro ao cadastrar: {str(e)}')

    turnos = Turno.objects.all()
    grupos = Grupo.objects.all()
    return render(request, 'escala/funcionario_form.html', {'turnos': turnos, 'grupos': grupos})


@login_required
def funcionario_editar(request, pk):
    """Edita um funcionário existente"""
    funcionario = get_object_or_404(Funcionario, pk=pk)
    
    if request.method == 'POST':
        try:
            funcionario.nome = request.POST.get('nome')
            funcionario.data_nascimento = request.POST.get('data_nascimento')
            funcionario.data_admissao = request.POST.get('data_admissao')
            funcionario.tipo = request.POST.get('tipo')
            if funcionario.tipo == 'FOLGUISTA':
                funcionario.turno = None
            else:
                turno_id = request.POST.get('turno')
                funcionario.turno = Turno.objects.get(id=turno_id) if turno_id else None
            grupo_id = request.POST.get('grupo')
            funcionario.grupo = Grupo.objects.get(id=grupo_id) if grupo_id else None
            funcionario.regime = request.POST.get('regime', '6x1')
            folga_fixa = request.POST.get('folga_fixa_dia')
            funcionario.folga_fixa_dia = int(folga_fixa) if folga_fixa not in ('', None) else None
            funcionario.ativo = request.POST.get('ativo') == 'on'
            funcionario.ferias_inicio = request.POST.get('ferias_inicio') or None
            funcionario.ferias_fim = request.POST.get('ferias_fim') or None
            funcionario.save()

            turnos_hab = request.POST.getlist('turnos_habilitados')
            grupos_hab = request.POST.getlist('grupos_habilitados')
            funcionario.turnos_habilitados.set(Turno.objects.filter(id__in=turnos_hab))
            funcionario.grupos_habilitados.set(Grupo.objects.filter(id__in=grupos_hab))

            messages.success(request, f'✅ Funcionário {funcionario.nome} atualizado com sucesso!')
            return redirect('escala:funcionario_lista')

        except Exception as e:
            messages.error(request, f'❌ Erro ao atualizar: {str(e)}')

    turnos = Turno.objects.all()
    grupos = Grupo.objects.all()
    context = {
        'funcionario': funcionario,
        'turnos': turnos,
        'grupos': grupos,
        'turnos_habilitados_ids': list(funcionario.turnos_habilitados.values_list('id', flat=True)),
        'grupos_habilitados_ids': list(funcionario.grupos_habilitados.values_list('id', flat=True)),
    }
    return render(request, 'escala/funcionario_form.html', context)


@login_required
@require_POST
def funcionario_deletar(request, pk):
    """Deleta um funcionário"""
    try:
        funcionario = get_object_or_404(Funcionario, pk=pk)
        nome = funcionario.nome
        funcionario.delete()
        
        messages.success(request, f'✅ Funcionário {nome} removido com sucesso!')
    except Exception as e:
        messages.error(request, f'❌ Erro ao remover: {str(e)}')
    
    return redirect('escala:funcionario_lista')


# ==================== TURNOS ====================

@login_required
def turno_lista(request):
    """Lista todos os turnos"""
    turnos = Turno.objects.all().order_by('horario_entrada')
    return render(request, 'escala/turno_lista.html', {'turnos': turnos})


@login_required
def turno_novo(request):
    """Cadastra novo turno"""
    if request.method == 'POST':
        try:
            Turno.objects.create(
                nome=request.POST.get('nome'),
                horario_entrada=request.POST.get('horario_entrada'),
                horario_saida=request.POST.get('horario_saida'),
                minimo_funcionarios=request.POST.get('minimo_funcionarios'),
            )
            messages.success(request, f'✅ Turno cadastrado!')
        except Exception as e:
            messages.error(request, f'❌ Erro: {str(e)}')
    
    return redirect('escala:turno_lista')


@login_required
def turno_editar(request, pk):
    """Edita turno"""
    turno = get_object_or_404(Turno, pk=pk)
    
    if request.method == 'POST':
        try:
            turno.nome = request.POST.get('nome')
            turno.horario_entrada = request.POST.get('horario_entrada')
            turno.horario_saida = request.POST.get('horario_saida')
            turno.minimo_funcionarios = request.POST.get('minimo_funcionarios')
            turno.save()
            messages.success(request, f'✅ Turno atualizado!')
        except Exception as e:
            messages.error(request, f'❌ Erro: {str(e)}')
    
    return redirect('escala:turno_lista')


@login_required
@require_POST
def turno_deletar(request, pk):
    """Deleta turno"""
    try:
        turno = get_object_or_404(Turno, pk=pk)
        turno.delete()
        messages.success(request, f'✅ Turno removido!')
    except Exception as e:
        messages.error(request, f'❌ Erro: {str(e)}')
    
    return redirect('escala:turno_lista')



@login_required
def exportar_escala_excel(request, pk):
    """Exporta escala para Excel — mesmo layout da tela: por setor, ordenado por turno, folguistas no final."""
    escala = get_object_or_404(Escala, pk=pk)
    dias_mes = monthrange(escala.ano, escala.mes)[1]

    # ── Cores e símbolos ──────────────────────────────────────────────────────
    CORES = {
        'TRABALHA': 'D4EDDA', 'FOLGA': 'F8D7DA', 'FOLGA_COMPENSADA': 'FFE0B2',
        'FALTA': 'EF9A9A', 'ATESTADO': 'E1BEE7', 'FERIAS': 'B3E5FC',
    }
    SIMBOLOS = {
        'TRABALHA': '✓', 'FOLGA': 'F', 'FOLGA_COMPENSADA': 'C',
        'FALTA': 'FT', 'ATESTADO': 'AF', 'FERIAS': 'FB',
    }
    ORDEM_TURNO = {'MANHA': 0, 'INTERMEDIARIO': 1, 'TARDE': 2, 'NOITE': 3}
    DIAS_SEM = {0: 'SEG', 1: 'TER', 2: 'QUA', 3: 'QUI', 4: 'SEX', 5: 'SAB', 6: 'DOM'}
    HEADER_CORES = ['F39C12', 'E74C3C', '34495E', '8E44AD', '16A085']

    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))

    # ── Carregar dados ────────────────────────────────────────────────────────
    dias_qs = list(DiaEscala.objects.filter(escala=escala).select_related(
        'funcionario', 'funcionario__grupo', 'funcionario__turno',
        'setor_coberto', 'turno_coberto'
    ))

    # Situação por funcionário e dia
    sit_map = {}       # {func_id: {dia: situacao}}
    cobertura_map = {} # {func_id: {dia: "SETOR/TUR"}}  — folguistas
    for d in dias_qs:
        fid = d.funcionario.id
        sit_map.setdefault(fid, {})[d.data.day] = d.situacao
        if d.funcionario.tipo == 'FOLGUISTA' and d.setor_coberto and d.turno_coberto:
            label = f"{d.setor_coberto.nome[:3]}/{d.turno_coberto.nome[0]}"
            cobertura_map.setdefault(fid, {})[d.data.day] = label

    # Regulares agrupados por setor
    regulares = Funcionario.objects.filter(tipo='REGULAR', ativo=True).select_related('grupo', 'turno')
    setores_map = {}  # {setor_nome: [func, ...]}
    for f in regulares:
        if f.grupo:
            setores_map.setdefault(f.grupo.nome, []).append(f)
    # Ordenar funcionários dentro de cada setor por turno
    for nome in setores_map:
        setores_map[nome].sort(key=lambda f: (ORDEM_TURNO.get(f.turno.nome.upper() if f.turno else '', 99), f.nome))

    # Folguistas
    folguistas = list(Funcionario.objects.filter(tipo='FOLGUISTA', ativo=True).order_by('nome'))

    # ── Criar workbook — uma aba única ───────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = f"Escala {escala.mes:02d}-{escala.ano}"

    # Título geral
    total_cols = dias_mes + 2  # nome + turno + dias
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    ws['A1'] = f'ESCALA {escala.mes:02d}/{escala.ano}'
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
    ws.row_dimensions[1].height = 28

    cur_row = 2  # linha atual de escrita

    def escrever_cabecalho_dias(row):
        ws.cell(row=row, column=1).value = 'FUNCIONÁRIO'
        ws.cell(row=row, column=2).value = 'TURNO'
        for c in [1, 2]:
            ws.cell(row=row, column=c).font = Font(bold=True, size=9)
            ws.cell(row=row, column=c).fill = PatternFill(start_color='BDC3C7', end_color='BDC3C7', fill_type='solid')
            ws.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
        for dia in range(1, dias_mes + 1):
            col = dia + 2
            d = date(escala.ano, escala.mes, dia)
            cell = ws.cell(row=row, column=col)
            cell.value = f"{dia}\n{DIAS_SEM[d.weekday()]}"
            cell.font = Font(bold=True, size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cor = 'F1C40F' if d.weekday() == 6 else 'BDC3C7'
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
        ws.row_dimensions[row].height = 28

    def escrever_funcionario(row, func, label_turno, dias_sit, dias_label=None):
        ws.cell(row=row, column=1).value = func.nome
        ws.cell(row=row, column=1).font = Font(bold=True, size=9)
        ws.cell(row=row, column=2).value = label_turno
        ws.cell(row=row, column=2).font = Font(size=8, color='444444')
        ws.cell(row=row, column=2).alignment = Alignment(horizontal='center', vertical='center')
        for dia in range(1, dias_mes + 1):
            col = dia + 2
            sit = dias_sit.get(dia, 'TRABALHA')
            cell = ws.cell(row=row, column=col)
            if dias_label:
                cell.value = dias_label.get(dia, SIMBOLOS.get(sit, '✓'))
            else:
                cell.value = SIMBOLOS.get(sit, '✓')
            cell.font = Font(size=8)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cor = CORES.get(sit, 'FFFFFF')
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type='solid')
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).border = thin

    # ── Seções por setor ─────────────────────────────────────────────────────
    for idx, (setor_nome, funcs) in enumerate(sorted(setores_map.items())):
        # Header do setor
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=total_cols)
        cell = ws.cell(row=cur_row, column=1)
        cell.value = f'🏢 {setor_nome.upper()}'
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cor_header = HEADER_CORES[idx % len(HEADER_CORES)]
        cell.fill = PatternFill(start_color=cor_header, end_color=cor_header, fill_type='solid')
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        # Cabeçalho de dias
        escrever_cabecalho_dias(cur_row)
        cur_row += 1

        # Funcionários
        for func in funcs:
            turno_label = func.turno.nome if func.turno else ''
            escrever_funcionario(cur_row, func, turno_label, sit_map.get(func.id, {}))
            cur_row += 1

        cur_row += 1  # Espaço entre setores

    # ── Seção folguistas ─────────────────────────────────────────────────────
    if folguistas:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=total_cols)
        cell = ws.cell(row=cur_row, column=1)
        cell.value = '🔄 FOLGUISTAS — cobertura de turnos'
        cell.font = Font(bold=True, size=11, color='FFFFFF')
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.fill = PatternFill(start_color='2980B9', end_color='2980B9', fill_type='solid')
        ws.row_dimensions[cur_row].height = 20
        cur_row += 1

        escrever_cabecalho_dias(cur_row)
        cur_row += 1

        for func in folguistas:
            dias_cob = cobertura_map.get(func.id, {})
            dias_sit = sit_map.get(func.id, {})
            # Para folguistas: mostrar cobertura quando TRABALHA, F quando FOLGA
            dias_label = {}
            for dia in range(1, dias_mes + 1):
                sit = dias_sit.get(dia, 'TRABALHA')
                if sit != 'TRABALHA':
                    dias_label[dia] = SIMBOLOS.get(sit, 'F')
                elif dia in dias_cob:
                    dias_label[dia] = dias_cob[dia]
                else:
                    dias_label[dia] = '✓'
            escrever_funcionario(cur_row, func, 'FOLGUISTA', dias_sit, dias_label)
            cur_row += 1

    # ── Largura das colunas ───────────────────────────────────────────────────
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 14
    for dia in range(1, dias_mes + 1):
        col_letter = ws.cell(row=3, column=dia + 2).column_letter
        ws.column_dimensions[col_letter].width = 7

    # Congelar painel após nome+turno
    ws.freeze_panes = 'C2'

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="Escala_{escala.mes:02d}_{escala.ano}.xlsx"'
    wb.save(response)
    return response


# ==================== CONFIGURAÇÕES ====================

@login_required
def configuracao_view(request):
    config = ConfiguracaoSistema.get()
    if request.method == 'POST':
        config.consecutivas_ativo = request.POST.get('consecutivas_ativo') == 'on'
        config.consecutivas_regime = request.POST.get('consecutivas_regime', 'AMBOS')
        config.domingo_ativo = request.POST.get('domingo_ativo') == 'on'
        config.save()
        messages.success(request, '✅ Configurações salvas!')
        return redirect('escala:configuracao')
    return render(request, 'escala/configuracao.html', {'config': config})


# ==================== GRUPOS ====================

@login_required
def grupo_lista(request):
    grupos = Grupo.objects.prefetch_related('turnos_operados__turno').all()
    return render(request, 'escala/grupo_lista.html', {'grupos': grupos})

@login_required
def grupo_novo(request):
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            grupo, _ = Grupo.objects.get_or_create(nome=nome)
            _salvar_setor_turnos(request, grupo)
            messages.success(request, f'✅ Grupo "{nome}" cadastrado!')
        return redirect('escala:grupo_lista')
    turnos = Turno.objects.all()
    return render(request, 'escala/grupo_form.html', {'turnos': turnos})

@login_required
def grupo_editar(request, pk):
    grupo = get_object_or_404(Grupo, pk=pk)
    if request.method == 'POST':
        nome = request.POST.get('nome', '').strip()
        if nome:
            grupo.nome = nome
            grupo.save()
            _salvar_setor_turnos(request, grupo)
            messages.success(request, f'✅ Grupo atualizado!')
        return redirect('escala:grupo_lista')
    turnos = Turno.objects.all()
    setor_turnos = {st.turno_id: st for st in SetorTurno.objects.filter(setor=grupo)}
    return render(request, 'escala/grupo_form.html', {
        'grupo': grupo,
        'turnos': turnos,
        'setor_turnos': setor_turnos,
    })


def _salvar_setor_turnos(request, grupo):
    """Lê turno_id[] e minimo_turno_<id>[] do POST e atualiza SetorTurno."""
    turno_ids = request.POST.getlist('turno_id[]')
    grupo.turnos_operados.all().delete()
    for tid in turno_ids:
        minimo = int(request.POST.get(f'minimo_turno_{tid}', 1) or 1)
        permite_zero = request.POST.get(f'permite_zero_{tid}') == 'on'
        try:
            turno = Turno.objects.get(id=int(tid))
            SetorTurno.objects.create(
                setor=grupo, turno=turno,
                minimo_funcionarios=max(1, minimo),
                permite_zero=permite_zero,
            )
        except (Turno.DoesNotExist, ValueError):
            pass

@login_required
@require_POST
def grupo_deletar(request, pk):
    grupo = get_object_or_404(Grupo, pk=pk)
    grupo.delete()
    messages.success(request, '✅ Grupo removido!')
    return redirect('escala:grupo_lista')


# ==================== FERIADOS ====================

@login_required
def feriado_lista(request):
    """Lista feriados do ano corrente (ou ano selecionado)"""
    ano_atual = date.today().year
    ano = int(request.GET.get('ano', ano_atual))

    feriados = Feriado.objects.filter(data__year=ano).order_by('data')
    anos_disponiveis = (
        Feriado.objects.dates('data', 'year').values_list('data__year', flat=True)
    )

    context = {
        'feriados': feriados,
        'ano': ano,
        'ano_atual': ano_atual,
        'anos_disponiveis': sorted(set(anos_disponiveis)),
        'total_feriados': feriados.count(),
    }
    return render(request, 'escala/feriado_lista.html', context)


@login_required
def feriado_novo(request):
    """Cadastra novo feriado"""
    if request.method == 'POST':
        try:
            nome = request.POST.get('nome')
            data = request.POST.get('data')
            tipo = request.POST.get('tipo')
            
            # Verificar se já existe feriado nesta data
            if Feriado.objects.filter(data=data).exists():
                messages.error(request, '❌ Já existe um feriado cadastrado nesta data!')
                return redirect('escala:feriado_novo')
            
            Feriado.objects.create(
                nome=nome,
                data=data,
                tipo=tipo
            )
            
            messages.success(request, f'✅ Feriado "{nome}" cadastrado com sucesso!')
            return redirect('escala:feriado_lista')
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao cadastrar: {str(e)}')
    
    return render(request, 'escala/feriado_form.html')


@login_required
def feriado_editar(request, pk):
    """Edita um feriado existente"""
    feriado = get_object_or_404(Feriado, pk=pk)
    
    if request.method == 'POST':
        try:
            feriado.nome = request.POST.get('nome')
            nova_data = request.POST.get('data')
            
            # Verificar se a nova data já existe (exceto o próprio feriado)
            if Feriado.objects.filter(data=nova_data).exclude(pk=pk).exists():
                messages.error(request, '❌ Já existe um feriado cadastrado nesta data!')
                return redirect('escala:feriado_editar', pk=pk)
            
            feriado.data = nova_data
            feriado.tipo = request.POST.get('tipo')
            feriado.save()
            
            messages.success(request, f'✅ Feriado "{feriado.nome}" atualizado!')
            return redirect('escala:feriado_lista')
            
        except Exception as e:
            messages.error(request, f'❌ Erro ao atualizar: {str(e)}')
    
    context = {'feriado': feriado}
    return render(request, 'escala/feriado_form.html', context)


@login_required
@require_POST
def feriado_deletar(request, pk):
    """Deleta um feriado"""
    try:
        feriado = get_object_or_404(Feriado, pk=pk)
        nome = feriado.nome
        feriado.delete()
        
        messages.success(request, f'✅ Feriado "{nome}" removido com sucesso!')
    except Exception as e:
        messages.error(request, f'❌ Erro ao remover: {str(e)}')
    
    return redirect('escala:feriado_lista')


# ==================== CALENDÁRIO ====================

@login_required
def calendario_view(request):
    """Calendário visual com escalas e ocorrências"""
    from calendar import monthrange
    from collections import defaultdict
    import json
    
    # Pegar mês/ano dos parâmetros ou usar atual
    hoje = date.today()
    mes = int(request.GET.get('mes', hoje.month))
    ano = int(request.GET.get('ano', hoje.year))
    
    # Validar mês
    if mes < 1 or mes > 12:
        mes = hoje.month
    
    # Buscar escala do mês
    try:
        escala = Escala.objects.get(mes=mes, ano=ano)
        dias_escala = DiaEscala.objects.filter(escala=escala).select_related(
            'funcionario', 'funcionario__turno', 'turno_coberto'
        )
    except Escala.DoesNotExist:
        escala = None
        dias_escala = []
    
    # Buscar feriados do mês
    data_inicio = date(ano, mes, 1)
    dias_mes = monthrange(ano, mes)[1]
    data_fim = date(ano, mes, dias_mes)
    
    feriados = Feriado.objects.filter(
        data__gte=data_inicio,
        data__lte=data_fim
    )
    feriados_dict = {f.data.day: f for f in feriados}

    # Buscar turnos reais do banco (ordenados por horário)
    turnos_db = list(Turno.objects.all().order_by('horario_entrada'))

    # Pré-indexar dias_escala por dia para evitar O(n²)
    dias_escala_por_dia = {}
    for dia_obj in dias_escala:
        d = dia_obj.data.day
        dias_escala_por_dia.setdefault(d, []).append(dia_obj)

    # Organizar dados por dia
    calendario_dias = []
    calendario_dias_json = []

    for dia in range(1, dias_mes + 1):
        data = date(ano, mes, dia)
        dia_semana = data.weekday()

        # Inicializa turnos_info com os nomes reais do banco
        turnos_info = {
            t.nome: {'nome': t.nome, 'funcionarios': [], 'minimo': t.minimo_funcionarios, 'atual': 0}
            for t in turnos_db
        }

        if escala:
            for dia_obj in dias_escala_por_dia.get(dia, []):
                if dia_obj.situacao != 'TRABALHA':
                    continue
                if dia_obj.funcionario.turno is None:
                    # Folguista: aparece no turno que está cobrindo
                    turno_obj = dia_obj.turno_coberto
                    if turno_obj is None:
                        continue
                    turno_nome = turno_obj.nome
                else:
                    turno_nome = dia_obj.funcionario.turno.nome
                if turno_nome in turnos_info:
                    turnos_info[turno_nome]['funcionarios'].append(dia_obj.funcionario.nome)

            for info in turnos_info.values():
                info['atual'] = len(info['funcionarios'])

        turnos_lista = list(turnos_info.values())
        feriado_obj = feriados_dict.get(dia)

        dia_info = {
            'dia': dia,
            'data': data,
            'dia_semana': dia_semana,
            'eh_domingo': dia_semana == 6,
            'eh_sabado': dia_semana == 5,
            'eh_hoje': data == hoje,
            'eh_passado': data < hoje,
            'eh_futuro': data > hoje,
            'feriado': feriado_obj,
            'turnos': turnos_info,
            'turnos_lista': turnos_lista,
        }

        calendario_dias.append(dia_info)

        dia_info_json = {
            'dia': dia,
            'data': data.isoformat(),
            'dia_semana': dia_semana,
            'eh_domingo': dia_semana == 6,
            'eh_sabado': dia_semana == 5,
            'eh_hoje': data == hoje,
            'eh_passado': data < hoje,
            'eh_futuro': data > hoje,
            'feriado': {
                'nome': feriado_obj.nome,
                'tipo': feriado_obj.get_tipo_display(),
                'eh_dia_util': feriado_obj.eh_dia_util()
            } if feriado_obj else None,
            'turnos_lista': turnos_lista,
        }

        calendario_dias_json.append(dia_info_json)
    
    # Calcular dias vazios no início do mês
    primeiro_dia = date(ano, mes, 1)
    dia_semana_inicial = primeiro_dia.weekday()
    # Ajustar para domingo = 0
    dias_vazios = (dia_semana_inicial + 1) % 7
    
    # Meses em português brasileiro
    meses_pt = [
        'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho',
        'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro'
    ]
    
    # Meses para navegação
    meses = [
        {'num': i, 'nome': meses_pt[i-1]}
        for i in range(1, 13)
    ]
    
    # Nome do mês atual em português
    mes_nome_pt = f"{meses_pt[mes-1]} {ano}"
    
    context = {
        'ano': ano,
        'mes': mes,
        'mes_nome': mes_nome_pt,
        'calendario_dias': calendario_dias,  # Para o template Django
        'calendario_dias_json': json.dumps(calendario_dias_json),  # Para JavaScript
        'dias_vazios': range(dias_vazios),  # Dias vazios no início
        'escala': escala,
        'meses': meses,
        'anos': range(2024, 2031),
    }
    
    return render(request, 'escala/calendario.html', context)